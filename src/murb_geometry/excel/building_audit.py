"""Building-level audit Excel workbook generation.

Reads persisted GeoParquet outputs and creates a filterable Excel workbook
with all classification, geometry, vertical-data, and quality fields exposed
at building level. Consumes pipeline outputs — no recalculation.
"""

import logging
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from murb_geometry import __version__
from murb_geometry.datastore import subset_path

logger = logging.getLogger(__name__)

# Columns to include in building audit (order matters)
AUDIT_COLUMNS = [
    # Source IDs
    "id",
    "source_id",
    "source",
    "_province",
    "_source_file",
    # Classification
    "confidence_level",
    "confidence_score",
    "rule_id",
    "rule_name",
    "evidence_fields",
    "reasoning",
    # Original attributes
    "type",
    "type_normalized",
    "units",
    "units_numeric",
    "floors",
    "floors_numeric",
    "height",
    "height_numeric",
    # Geometry metrics
    "footprint_area_m2",
    "perimeter_m",
    "mrr_length_m",
    "mrr_width_m",
    "mrr_area_m2",
    "aspect_ratio",
    "orientation_deg",
    "compactness",
    "rectangularity",
    "convexity",
    "hole_count",
    "hole_area_m2",
    "vertex_count",
    # Location
    "csduid",
    "csdname",
    "prov_terr",
]

# Suspicious-record filter conditions
SUSPICIOUS_RULES = [
    ("area_extreme_small", "footprint_area_m2", "<", 50),
    ("area_extreme_large", "footprint_area_m2", ">", 20000),
    ("aspect_extreme", "aspect_ratio", ">", 10),
    ("compactness_extreme_low", "compactness", "<", 0.1),
]


def create_building_audit_workbook(
    output_path: Path,
    precision_path: Path | None = None,
    tiered_path: Path | None = None,
) -> Path:
    """Create a building-level audit Excel workbook from persisted outputs.

    Reads GeoParquet files — no recalculation logic.

    Parameters
    ----------
    output_path
        Path to write the .xlsx file.
    precision_path
        Path to precision-pathway GeoParquet.
    tiered_path
        Path to tiered-pathway GeoParquet.

    Returns
    -------
    Path
        Path to created workbook.
    """
    precision_path = precision_path or subset_path("precision")
    tiered_path = tiered_path or subset_path("tiered")

    wb = Workbook()

    # --- Read Me ---
    ws = wb.active
    assert ws is not None
    ws.title = "Read Me"
    _write_readme(ws, precision_path, tiered_path)

    # --- Precision Pathway ---
    if precision_path.exists():
        gdf_p = gpd.read_parquet(precision_path)
        _write_building_sheet(wb, gdf_p, "Precision Buildings")
        _write_suspicious_sheet(wb, gdf_p, "Precision Suspicious")
        _write_summary_sheet(wb, gdf_p, "Precision Summary")
        logger.info("Precision sheet: %d buildings", len(gdf_p))

    # --- Tiered Pathway ---
    if tiered_path.exists():
        gdf_t = gpd.read_parquet(tiered_path)
        _write_building_sheet(wb, gdf_t, "Tiered Buildings")
        _write_suspicious_sheet(wb, gdf_t, "Tiered Suspicious")
        _write_summary_sheet(wb, gdf_t, "Tiered Summary")
        logger.info("Tiered sheet: %d buildings", len(gdf_t))

    # --- Field Dictionary ---
    _write_field_dictionary(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Building audit workbook: %s", output_path)
    return output_path


def _write_readme(
    ws: Any,
    precision_path: Path,
    tiered_path: Path,
) -> None:
    """Write Read Me sheet with metadata."""
    header_font = Font(bold=True, size=14)
    ws["A1"] = "MURB Building-Level Audit Workbook"
    ws["A1"].font = header_font

    info = [
        ("Generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ("Software", f"murb-geometry v{__version__}"),
        ("Source", "Statistics Canada Open Database of Buildings v3"),
        ("CRS", "EPSG:3347 (NAD83 / Statistics Canada Lambert)"),
        ("Pathway", "Option C - Multi-pathway (precision + tiered)"),
        ("Precision data", str(precision_path)),
        ("Tiered data", str(tiered_path)),
        ("", ""),
        ("Sheets", ""),
        ("Precision Buildings", "All precision-pathway MURBs with full metrics"),
        ("Precision Suspicious", "Records with extreme/unusual values for review"),
        ("Precision Summary", "Aggregate statistics"),
        ("Tiered Buildings", "All tiered-pathway MURBs with full metrics"),
        ("Tiered Suspicious", "Records with extreme/unusual values for review"),
        ("Tiered Summary", "Aggregate statistics"),
        ("Field Dictionary", "Column definitions and units"),
    ]
    for i, (key, val) in enumerate(info, start=3):
        ws[f"A{i}"] = key
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = val

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


def _write_building_sheet(
    wb: Workbook,
    gdf: gpd.GeoDataFrame,
    sheet_name: str,
) -> None:
    """Write building-level data with frozen header and auto-filter."""
    # Select available columns
    cols = [c for c in AUDIT_COLUMNS if c in gdf.columns]
    df = gdf[cols].copy()

    # Drop geometry column if present (not useful in Excel)
    if "geometry" in df.columns:
        df = df.drop(columns=["geometry"])

    ws = wb.create_sheet(sheet_name)

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            # Convert numpy types for Excel compatibility
            if pd.isna(val):
                val = None
            elif hasattr(val, "item"):
                val = val.item()
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # Auto-width columns
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(
            max(len(str(df.columns[col_idx - 1])) + 2, 12), 30
        )


def _write_suspicious_sheet(
    wb: Workbook,
    gdf: gpd.GeoDataFrame,
    sheet_name: str,
) -> None:
    """Write suspicious records filtered by extreme values."""
    suspicious_mask = pd.Series(False, index=gdf.index)
    reasons: list[str] = [""] * len(gdf)

    for rule_name, field, op, threshold in SUSPICIOUS_RULES:
        if field not in gdf.columns:
            continue
        if op == "<":
            mask = gdf[field] < threshold
        elif op == ">":
            mask = gdf[field] > threshold
        else:
            continue

        for idx in gdf.index[mask]:
            pos = gdf.index.get_loc(idx)
            if reasons[pos]:
                reasons[pos] += f"; {rule_name}"
            else:
                reasons[pos] = rule_name

        suspicious_mask = suspicious_mask | mask

    suspicious = gdf[suspicious_mask].copy()
    suspicious["suspicious_reason"] = [
        r for r, m in zip(reasons, suspicious_mask, strict=False) if m
    ]

    if suspicious.empty:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = "No suspicious records found."
        return

    cols = [c for c in AUDIT_COLUMNS if c in suspicious.columns] + ["suspicious_reason"]
    _write_building_sheet.__wrapped__ if hasattr(_write_building_sheet, "__wrapped__") else None

    ws = wb.create_sheet(sheet_name)
    df = suspicious[cols].copy()
    if "geometry" in df.columns:
        df = df.drop(columns=["geometry"])

    # Header
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            elif hasattr(val, "item"):
                val = val.item()
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    logger.info("  %s: %d suspicious records", sheet_name, len(suspicious))


def _write_summary_sheet(
    wb: Workbook,
    gdf: gpd.GeoDataFrame,
    sheet_name: str,
) -> None:
    """Write aggregate summary statistics."""
    ws = wb.create_sheet(sheet_name)

    metrics = [
        "footprint_area_m2",
        "aspect_ratio",
        "compactness",
        "rectangularity",
        "convexity",
        "mrr_length_m",
        "mrr_width_m",
        "perimeter_m",
    ]

    # Header
    headers = ["Metric", "N", "Missing", "Min", "P25", "Median", "P75", "Max", "Mean", "Std"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    row_idx = 2
    for metric in metrics:
        if metric not in gdf.columns:
            continue
        values = gdf[metric].dropna()
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=len(values))
        ws.cell(row=row_idx, column=3, value=len(gdf) - len(values))
        if len(values) > 0:
            ws.cell(row=row_idx, column=4, value=round(float(values.min()), 2))
            ws.cell(row=row_idx, column=5, value=round(float(values.quantile(0.25)), 2))
            ws.cell(row=row_idx, column=6, value=round(float(values.median()), 2))
            ws.cell(row=row_idx, column=7, value=round(float(values.quantile(0.75)), 2))
            ws.cell(row=row_idx, column=8, value=round(float(values.max()), 2))
            ws.cell(row=row_idx, column=9, value=round(float(values.mean()), 2))
            ws.cell(row=row_idx, column=10, value=round(float(values.std()), 2))
        row_idx += 1

    # Classification breakdown
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Classification Breakdown")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    if "confidence_level" in gdf.columns:
        for level, count in gdf["confidence_level"].value_counts().items():
            ws.cell(row=row_idx, column=1, value=str(level))
            ws.cell(row=row_idx, column=2, value=int(count))
            row_idx += 1


def _write_field_dictionary(wb: Workbook) -> None:
    """Write field definitions sheet."""
    ws = wb.create_sheet("Field Dictionary")

    fields = [
        ("id", "Source building ID", "TEXT"),
        ("source_id", "Source-specific record ID", "TEXT"),
        ("source", "Data source organization", "TEXT"),
        ("_province", "Province/territory code", "TEXT"),
        ("confidence_level", "Classification confidence class", "TEXT"),
        ("confidence_score", "Confidence score (0.0-1.0)", "FLOAT"),
        ("rule_id", "Classification rule that fired", "TEXT"),
        ("type_normalized", "Normalized building type", "TEXT"),
        ("units_numeric", "Parsed dwelling unit count", "INT"),
        ("floors_numeric", "Parsed storey count", "INT"),
        ("height_numeric", "Parsed building height (m)", "FLOAT"),
        ("footprint_area_m2", "Ground footprint area", "FLOAT (m2)"),
        ("perimeter_m", "Footprint perimeter", "FLOAT (m)"),
        ("mrr_length_m", "MRR major axis length", "FLOAT (m)"),
        ("mrr_width_m", "MRR minor axis length", "FLOAT (m)"),
        ("aspect_ratio", "Length/width (always >= 1)", "FLOAT"),
        ("orientation_deg", "Major axis azimuth from N", "FLOAT (deg)"),
        ("compactness", "Polsby-Popper: 4piA/P^2", "FLOAT [0,1]"),
        ("rectangularity", "Area / MRR area", "FLOAT [0,1]"),
        ("convexity", "Area / convex hull area", "FLOAT [0,1]"),
        ("hole_count", "Number of interior holes", "INT"),
        ("vertex_count", "Polygon vertex count", "INT"),
    ]

    headers = ["Field", "Description", "Type/Unit"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    for row_idx, (field, desc, dtype) in enumerate(fields, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=desc)
        ws.cell(row=row_idx, column=3, value=dtype)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
