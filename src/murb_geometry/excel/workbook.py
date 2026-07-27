"""Excel workbook generation for MURB geometry reports.

Creates formatted .xlsx workbooks with standard sheets:
- Read Me: report metadata and assumptions
- Data Quality: completeness by province and field
- MURB Summary: descriptive statistics
- Field Dictionary: field definitions
"""

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from murb_geometry import __version__


def create_summary_workbook(
    output_path: Path,
    completeness_data: list[dict[str, Any]] | None = None,
    summary_stats: list[dict[str, Any]] | None = None,
    metadata: dict[str, str] | None = None,
) -> Path:
    """Create a formatted Excel summary workbook.

    Parameters
    ----------
    output_path
        Path to write the .xlsx file.
    completeness_data
        List of dicts with province/field completeness percentages.
    summary_stats
        List of descriptive statistics dictionaries.
    metadata
        Report metadata (filters, date, version, etc.).

    Returns
    -------
    Path
        Path to the created workbook.
    """
    wb = Workbook()

    # --- Read Me sheet ---
    ws_readme = wb.active
    assert ws_readme is not None
    ws_readme.title = "Read Me"
    _write_readme_sheet(ws_readme, metadata)

    # --- Data Quality sheet ---
    if completeness_data:
        ws_quality = wb.create_sheet("Data Quality")
        _write_completeness_sheet(ws_quality, completeness_data)

    # --- MURB Summary sheet ---
    if summary_stats:
        ws_summary = wb.create_sheet("MURB Summary")
        _write_summary_sheet(ws_summary, summary_stats)

    # --- Field Dictionary sheet ---
    ws_dict = wb.create_sheet("Field Dictionary")
    _write_dictionary_sheet(ws_dict)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_readme_sheet(ws: Any, metadata: dict[str, str] | None) -> None:
    """Write the Read Me sheet with report metadata."""
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)

    ws["A1"] = "Canadian MURB Geometry Analysis — Summary Report"
    ws["A1"].font = header_font

    rows = [
        ("Report generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ("Software version", __version__),
        ("Data source", "Statistics Canada Open Database of Buildings v3"),
        ("Licence", "Open Government Licence — Canada"),
        ("CRS", "EPSG:3347 (NAD83 / Statistics Canada Lambert)"),
    ]
    if metadata:
        for key, value in metadata.items():
            rows.append((key, value))

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


def _write_completeness_sheet(ws: Any, completeness_data: list[dict[str, Any]]) -> None:
    """Write data quality / completeness matrix."""
    if not completeness_data:
        return

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF")

    # Write headers from first record's keys
    headers = list(completeness_data[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, record in enumerate(completeness_data, 2):
        for col_idx, h in enumerate(headers, 1):
            value = record.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Conditional formatting for percentages
            if (isinstance(value, (int, float)) and "%" in h.lower()) or "pct" in h.lower():
                if value >= 50:
                    cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                elif value >= 10:
                    cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(ws: Any, summary_stats: list[dict[str, Any]]) -> None:
    """Write descriptive statistics."""
    if not summary_stats:
        return

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF")

    headers = list(summary_stats[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill

    for row_idx, record in enumerate(summary_stats, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(h))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "B2"


def _write_dictionary_sheet(ws: Any) -> None:
    """Write field definitions."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF")

    headers = ["Field", "Type", "Unit", "Description", "Source"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill

    fields = [
        ("footprint_area_m2", "float", "m2", "Building footprint area", "Calculated"),
        ("perimeter_m", "float", "m", "Footprint perimeter", "Calculated"),
        ("mrr_length_m", "float", "m", "Minimum rotated rectangle major axis", "Calculated"),
        ("mrr_width_m", "float", "m", "Minimum rotated rectangle minor axis", "Calculated"),
        ("aspect_ratio", "float", "-", "Length / width ratio", "Calculated"),
        ("orientation_deg", "float", "degrees", "Major axis azimuth from north", "Calculated"),
        ("compactness", "float", "-", "Polsby-Popper: 4pi*A/P^2", "Calculated"),
        ("rectangularity", "float", "-", "Area / MRR area", "Calculated"),
        ("convexity", "float", "-", "Area / convex hull area", "Calculated"),
        ("floors", "int", "-", "Number of storeys", "Observed/Enriched"),
        ("height", "float", "m", "Building height", "Observed/Enriched"),
        ("units", "int", "-", "Number of dwelling units", "Observed/Enriched"),
        ("type_normalized", "str", "-", "Normalized building type", "Derived"),
        ("classification", "str", "-", "MURB confidence level", "Derived"),
    ]

    for row_idx, field_row in enumerate(fields, 2):
        for col_idx, value in enumerate(field_row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A2"
