"""Unit tests for Excel workbook generation."""

from pathlib import Path

from openpyxl import load_workbook

from murb_geometry.excel.workbook import create_summary_workbook


def test_create_empty_workbook(tmp_path: Path) -> None:
    """Workbook is created with default sheets."""
    output = tmp_path / "test_report.xlsx"
    result = create_summary_workbook(output)
    assert result == output
    assert output.exists()

    wb = load_workbook(output)
    assert "Read Me" in wb.sheetnames
    assert "Field Dictionary" in wb.sheetnames


def test_workbook_with_completeness(tmp_path: Path) -> None:
    """Workbook includes Data Quality sheet when data provided."""
    output = tmp_path / "test_report.xlsx"
    completeness = [
        {"province": "NS", "type_pct": 25.0, "floors_pct": 0.8, "units_pct": 23.0},
        {"province": "ON", "type_pct": 14.4, "floors_pct": 5.9, "units_pct": 1.0},
    ]
    create_summary_workbook(output, completeness_data=completeness)

    wb = load_workbook(output)
    assert "Data Quality" in wb.sheetnames
    ws = wb["Data Quality"]
    assert ws.cell(row=1, column=1).value == "province"
    assert ws.cell(row=2, column=1).value == "NS"


def test_workbook_with_summary_stats(tmp_path: Path) -> None:
    """Workbook includes MURB Summary sheet when stats provided."""
    output = tmp_path / "test_report.xlsx"
    stats = [
        {"field": "footprint_area_m2", "count": 1000, "mean": 450.5, "median": 380.0},
        {"field": "aspect_ratio", "count": 1000, "mean": 2.1, "median": 1.8},
    ]
    create_summary_workbook(output, summary_stats=stats)

    wb = load_workbook(output)
    assert "MURB Summary" in wb.sheetnames


def test_workbook_with_metadata(tmp_path: Path) -> None:
    """Workbook Read Me includes custom metadata."""
    output = tmp_path / "test_report.xlsx"
    metadata = {"Province filter": "NS", "Min area": "200 m2"}
    create_summary_workbook(output, metadata=metadata)

    wb = load_workbook(output)
    ws = wb["Read Me"]
    # Check that metadata appears somewhere in the sheet
    values = [ws.cell(row=r, column=1).value for r in range(1, 15)]
    assert "Province filter" in values
